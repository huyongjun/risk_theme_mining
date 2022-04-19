import xlrd
import re
import openpyxl

def cal_imp():
    # 打开Excel文件
    wb = xlrd.open_workbook('E:\\研究生\\项目\\群组行为突发事件主题发现模型\\风险+心智图谱\\文心系统心理分析\中文LIWC分析_m\\计算指标.xlsx')
    sheet1 = wb.sheet_by_name('相关关系值表_cal')#通过excel工作表名称(rank)获取工作表
    sheet2 = wb.sheet_by_name('各词功能词类别表')  # 通过excel工作表名称(rank)获取工作表
    datadic_rel = {}  #创建相关关系表列表
    datalist_class = [] #创建各词功能词类别列表
    datadic_word_imp={} #创建各词重要程度表列表
    #获取相关关系表数据，放入list
    for a in range(1, sheet1.nrows):
        rellist = []
        cells1 = sheet1.row_values(a)# 每行数据赋值给cells
        cells1[1] = cells1[1].encode('utf-8').decode('utf-8-sig')
        cells1[2] = cells1[2].encode('utf-8').decode('utf-8-sig')
        rellist.append(cells1[1])
        rellist.append(cells1[2])
        rellist.append(cells1[3])
        rellist.append(cells1[4])
        rellist.append(cells1[5])
        rellist.append(cells1[6])
        rellist.append(cells1[7])
        #把相关关系整理为字典
        datadic_rel[cells1[0]]=rellist

    #获取各词功能词类别表数据，放入list
    for b in range(1, sheet2.nrows):
        cells2 = sheet2.row_values(b)# 每行数据赋值给cells
        #rel_count=len(cells)-1#计数每个词属于多少个功能类别
        # 解决字符串转列表 list 出现\ufeff
        cells2[0] = cells2[0].encode('utf-8').decode('utf-8-sig')
        datalist_class.append(cells2)

    ###计算各词在五人格维度下的重要程度BFM
    #遍历每个词
    for wordlist in datalist_class:
        implist=[]
        # 遍历每个人格维度
        for i in range(3, 8):
            sum = 0
            class_count = 0
            #遍历每个分类
            for j in range(1,len(wordlist)):
                if wordlist[j]=='':
                    break
                #计算各人格特征下的重要程度
                sum=sum+datadic_rel[wordlist[j]][i-1]
                if datadic_rel[wordlist[j]][i-1]!=0:
                    class_count=class_count+1
            if class_count!=0:
                imp=sum/class_count
            else:imp=0
            implist.append(imp)
        datadic_word_imp[wordlist[0]]=implist

    #数据写入excel,xlsx格式
    data = open('E:\\研究生\\项目\\群组行为突发事件主题发现模型\\风险+心智图谱\\文心系统心理分析\中文LIWC分析_m\\计算指标_result.xlsx', 'r')
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet

    # 注意：'cell'函数中行列起始值为1
    i = 1
    for key in datadic_word_imp.keys():
        j = 1
        outws.cell(column=j, row=i, value="%s" % key)
        for k in range(len(datadic_word_imp[key])):
            j=j+1
            outws.cell(column=j, row=i, value="%s" % datadic_word_imp[key][k])
        i = i + 1
    savexlsx = 'E:\\研究生\\项目\\群组行为突发事件主题发现模型\\风险+心智图谱\\文心系统心理分析\中文LIWC分析_m\\计算指标_result.xlsx'
    outwb.save(savexlsx)  # 保存结果
    data.close()

def main():
    cal_imp()
    print("计算顺利结束！！")

main()