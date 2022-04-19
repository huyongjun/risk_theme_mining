from openpyxl import load_workbook

## 处理微博id（导入excel只保留25位）
def pro_weiboid():
    weibos = []  #TODO:weibo列表
    ##---wm 读取weibo列表
    file_home = r'E:\研究生\项目\群组行为突发事件主题发现模型\风险+心智图谱\爬数据\标签数据\3labels\带预测标签4w数据.xlsx'
    file_save=r'E:\研究生\项目\群组行为突发事件主题发现模型\风险+心智图谱\爬数据\标签数据\3labels\带预测标签4w数据_r.xlsx'
    wb = load_workbook(filename=file_home)
    sheet1 = wb.worksheets[0]
    # ws = wb['hot_weibo'] # 根据Sheet1这个sheet名字来获取该sheet

    # 获取表格所有行和列，两者都是可迭代的
    rows = sheet1.rows
    # columns = sheet1.columns
    # line = [col.value for col in row] ##遍历每一列
    # 迭代所有的行
    for row in rows:
        comment_id = str(row[7].value)
        if comment_id == 'comment_id':
            continue
        cindex=comment_id.find('c')
        weibo_id=str(comment_id[0:cindex])+'w'
        sheet1.cell(row[0].row, 2).value = weibo_id
        print(weibo_id)

    wb.save(file_save)  # 保存修改后的excel


def test():
    ne_count = 5
    n_count = 3
    p_count = 0
    maxp = max(ne_count, p_count, n_count)
    weibo_label = '0' if maxp == ne_count else ('1' if maxp == p_count else '2')
    print(weibo_label)

def main():
    # pro_weiboid()
    test()
main()