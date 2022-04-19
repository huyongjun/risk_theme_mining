import requests
from lxml import etree
from urllib import parse
import re
import configparser
import json
import pymysql.cursors
import re
from tools.Mysql_Process import mysqlHelper
from tools.Mysql_Process import get_db
import time
import random
import xlrd
from bs4 import BeautifulSoup
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import math


####主要用excel处理数据实现微博、话题感情统计
def proby_excel():
    weibos = []  #TODO:weibo列表
    ##---wm 读取weibo列表
    file_home = r'E:\研究生\项目\群组行为突发事件主题发现模型\风险+心智图谱\爬数据\评论情感+微博情感数据处理\hot_weibo_topic.xlsx'
    file_save=r'E:\研究生\项目\群组行为突发事件主题发现模型\风险+心智图谱\爬数据\评论情感+微博情感数据处理\hot_weibo_labeled.xlsx'
    wb = load_workbook(filename=file_home)
    sheet1 = wb.worksheets[0]
    # ws = wb['hot_weibo'] # 根据Sheet1这个sheet名字来获取该sheet

    # 获取表格所有行和列，两者都是可迭代的
    rows = sheet1.rows
    # columns = sheet1.columns
    # line = [col.value for col in row] ##遍历每一列

    ##连接mysql数据库
    try:
        mh = mysqlHelper(get_db()[0], get_db()[1], get_db()[2], get_db()[3], get_db()[4], int(get_db()[5]))
        # mh.open()
    except Exception as e:
            print(e)
    # 迭代所有的行
    for row in rows:
        weibo_id=str(row[1].value)
        ne_count=0
        n_count=0
        p_count=0
        weibo_label=0
        if weibo_id=='weibo_id':
            continue
        #查询话题序号
        select_sql = "SELECT * FROM weibo_4wcomment_labeled WHERE weibo_id = %s"
        result=list(mh.findAll(select_sql,weibo_id))
        for comment in result:
            if str(comment[13])=='0':
                ne_count=ne_count + 1
            elif str(comment[13])=='1':
                p_count= p_count + 1
            elif str(comment[13])=='2':
                n_count=n_count +1
            weibo_label=max(ne_count,p_count,n_count)
        sheet1.cell(row[0].row, 11).value = ne_count
        sheet1.cell(row[0].row, 12).value = p_count
        sheet1.cell(row[0].row, 13).value = n_count
        sheet1.cell(row[0].row, 14).value = weibo_label
    # mh.close()

    wb.save(file_save) # 保存修改后的excel


####主要用mysql处理数据实现【微博】感情计数统计【未考虑点赞数】【未考虑马太效应】
def proby_mysql():
    ##连接mysql数据库
    try:
        mh = mysqlHelper(get_db()[0], get_db()[1], get_db()[2], get_db()[3], get_db()[4], int(get_db()[5]))
        # mh.open()
    except Exception as e:
        print(e)
    select_sql = "SELECT * FROM hot_weibo_topic"
    weibo_list = list(mh.select(select_sql))
    for w in weibo_list:
        weibo_id=w[1]
        ne_count = 0
        n_count = 0
        p_count = 0
        weibo_label = 0
        # 查询话题序号
        select_sql = "SELECT * FROM weibo_4wcomment_labeled WHERE weibo_id = %s"
        result = list(mh.findAll(select_sql, weibo_id))
        for comment in result:
            if str(comment[13]) == '0':
                ne_count = ne_count + 1
            elif str(comment[13]) == '1':
                p_count = p_count + 1
            elif str(comment[13]) == '2':
                n_count = n_count + 1
        maxp = max(ne_count, p_count, n_count)
        weibo_label = '0' if maxp == ne_count else ('1' if maxp == p_count else '2')
        ##更新数据
        update_sql="UPDATE hot_weibo_topic SET ne_count=%s,p_count=%s,n_count=%s,weibo_label=%s WHERE weibo_id=%s"
        result=mh.cud(update_sql,(str(ne_count),str(p_count),str(n_count),str(weibo_label),weibo_id))

        ##提交数据
        mh.tijiao();

    mh.close()


####主要用mysql处理数据【评论】感情处理统计，【考虑点赞数】【考虑马太效应】
def pro_commentemo_byms_withnoME():
    ##连接mysql数据库
    try:
        mh = mysqlHelper(get_db()[0], get_db()[1], get_db()[2], get_db()[3], get_db()[4], int(get_db()[5]))
        # mh.open()
    except Exception as e:
        print(e)
    select_sql = "SELECT * FROM weibo_4wcomment_labeled"
    comment_list = list(mh.select(select_sql))
    for c in comment_list:
        id=c[0]
        pro_count=0
        like_count=int(c[5])
        if like_count==0:
            pro_count=math.log(like_count+1)
        else:
            for i in range(1,like_count+1):
                pro_count=pro_count+math.log(i+1)
        pro_count=('%.3f' %pro_count)
        ##更新数据
        update_sql="UPDATE weibo_4wcomment_labeled SET pro_likecount=%s where id=%s"
        result=mh.cud(update_sql,(str(pro_count),id))

        ##提交数据
        mh.tijiao();

    mh.close()

    print()


####主要用mysql处理数据实现【微博】感情计数统计【考虑点赞数】【考虑马太效应】
def pro_weiboemo_byms_withnoME():
    ##连接mysql数据库
    try:
        mh = mysqlHelper(get_db()[0], get_db()[1], get_db()[2], get_db()[3], get_db()[4], int(get_db()[5]))
        # mh.open()
    except Exception as e:
        print(e)
    select_sql = "SELECT * FROM hot_weibo_topic"
    weibo_list = list(mh.select(select_sql))
    for w in weibo_list:
        weibo_id = w[1]
        pro_ne_count = 0.000
        pro_n_count = 0.000
        pro_p_count = 0.000
        pro_weibo_label = 0.000
        # 查询话题序号
        select_sql = "SELECT * FROM weibo_4wcomment_labeled WHERE weibo_id = %s"
        result = list(mh.findAll(select_sql, weibo_id))
        for comment in result:
            if str(comment[14]) == '0':
                pro_ne_count = pro_ne_count + float(comment[6])
            elif str(comment[14]) == '1':
                pro_p_count = pro_p_count + float(comment[6])
            elif str(comment[14]) == '2':
                pro_n_count = pro_n_count + float(comment[6])
        maxp = max(pro_ne_count, pro_p_count, pro_n_count)
        pro_weibo_label = '0' if maxp == pro_ne_count else ('1' if maxp == pro_p_count else '2')
        ##更新数据
        update_sql = "UPDATE hot_weibo_topic SET pro_ne_count=%s,pro_p_count=%s,pro_n_count=%s,pro_weibo_label=%s WHERE weibo_id=%s"
        result = mh.cud(update_sql, (str(pro_ne_count), str(pro_p_count), str(pro_n_count), str(pro_weibo_label), weibo_id))

        ##提交数据
        mh.tijiao();

    mh.close()


####主要用mysql处理数据【评论】感情处理统计，【考虑点赞数】【未考虑马太效应】
###为了测试，考虑点赞数时，但未考虑马太效应时，weibo的标签会有多少不同
def commentemo_byms_withnoME():
    ##连接mysql数据库
    try:
        mh = mysqlHelper(get_db()[0], get_db()[1], get_db()[2], get_db()[3], get_db()[4], int(get_db()[5]))
        # mh.open()
    except Exception as e:
        print(e)
    select_sql = "SELECT * FROM hot_weibo_topic"
    weibo_list = list(mh.select(select_sql))
    diff_count=0
    for w in weibo_list:
        weibo_id = w[1]
        pro_ne_count = 0
        pro_n_count = 0
        pro_p_count = 0
        pro_weibo_label = 0
        # 查询话题序号
        select_sql = "SELECT * FROM weibo_4wcomment_labeled WHERE weibo_id = %s"
        result = list(mh.findAll(select_sql, weibo_id))
        for comment in result:
            if str(comment[14]) == '0':
                pro_ne_count = pro_ne_count + comment[5]+1
            elif str(comment[14]) == '1':
                pro_p_count = pro_p_count + comment[5]+1
            elif str(comment[14]) == '2':
                pro_n_count = pro_n_count + comment[5]+1
        maxp = max(pro_ne_count, pro_p_count, pro_n_count)
        pro_weibo_label = '0' if maxp == pro_ne_count else ('1' if maxp == pro_p_count else '2')
        if pro_weibo_label!=str(w[17]):
            diff_count=diff_count+1

    print(diff_count)

def main():
    # proby_excel()
    # proby_mysql()
    # pro_commentemo_byms_withnoME()
    # pro_weiboemo_byms_withnoME()
    commentemo_byms_withnoME()



main()